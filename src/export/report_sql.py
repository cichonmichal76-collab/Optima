from __future__ import annotations

import math
import re
import subprocess
from datetime import datetime
from html import escape
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def export_sql_report_xlsx(
    file_path: Path,
    *,
    title: str,
    headers: list[str],
    rows: list[dict[str, object]],
    notes: Iterable[str] = (),
    filter_label: str = "",
    include_chart: bool = False,
) -> Path:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Raport"
    summary_sheet.append(["Pole", "Wartość"])
    summary_sheet.append(["Tytuł", title])
    summary_sheet.append(["Wygenerowano", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary_sheet.append(["Filtr", filter_label or "bez ograniczenia dat"])
    summary_sheet.append(["Liczba wierszy", len(rows)])
    for note in notes:
        summary_sheet.append(["Nota", str(note)])
    _style_summary_sheet(summary_sheet)

    data_sheet = workbook.create_sheet("Dane")
    if headers:
        data_sheet.append(headers)
        for row in rows:
            data_sheet.append([row.get(header, "") for header in headers])
        _style_data_sheet(data_sheet, headers, rows)
    else:
        data_sheet.append(["Brak danych"])

    if include_chart:
        chart_points = build_chart_points(headers, rows)
        if chart_points:
            _append_chart_sheet(workbook, chart_points, title)

    workbook.save(file_path)
    return file_path


def export_sql_report_pdf(
    file_path: Path,
    *,
    title: str,
    headers: list[str],
    rows: list[dict[str, object]],
    notes: Iterable[str] = (),
    filter_label: str = "",
    include_chart: bool = False,
    browser_path: str | None = None,
) -> Path:
    html = build_sql_report_html(
        title=title,
        headers=headers,
        rows=rows,
        notes=notes,
        filter_label=filter_label,
        include_chart=include_chart,
    )
    _render_pdf_via_browser(file_path, html, browser_path=browser_path)
    return file_path


def build_sql_report_html(
    *,
    title: str,
    headers: list[str],
    rows: list[dict[str, object]],
    notes: Iterable[str] = (),
    filter_label: str = "",
    include_chart: bool = False,
) -> str:
    notes_html = "".join(f"<li>{escape(str(note))}</li>" for note in notes)
    rows_html = "".join(
        "<tr>"
        + "".join(f"<td>{escape(str(row.get(header, '')))}</td>" for header in headers)
        + "</tr>"
        for row in rows
    )
    chart_html = ""
    if include_chart:
        chart_points = build_chart_points(headers, rows)
        if chart_points:
            chart_html = _build_chart_svg(chart_points)

    return f"""
<!doctype html>
<html lang="pl">
<head>
  <meta charset="utf-8">
  <title>{escape(title)}</title>
  <style>
    body {{ font-family: "Segoe UI", Arial, sans-serif; margin: 24px; color: #17201c; }}
    h1, h2 {{ margin: 0 0 12px; }}
    .meta {{ margin: 16px 0 20px; padding: 16px; border: 1px solid #d7ded8; border-radius: 12px; background: #f7f8f6; }}
    .meta strong {{ color: #176b5b; }}
    ul {{ margin: 12px 0 0; padding-left: 18px; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 18px; }}
    th, td {{ border: 1px solid #d7ded8; padding: 8px 10px; text-align: left; vertical-align: top; font-size: 12px; }}
    th {{ background: #e7f8f1; }}
    .muted {{ color: #66736e; }}
    .empty {{ padding: 18px; border: 1px dashed #d7ded8; border-radius: 12px; background: #f7f8f6; }}
    .chart {{ margin-top: 20px; padding: 16px; border: 1px solid #d7ded8; border-radius: 12px; background: #ffffff; }}
    .footer {{ margin-top: 18px; color: #66736e; font-size: 11px; }}
    svg text {{ font-family: "Segoe UI", Arial, sans-serif; }}
  </style>
</head>
<body>
  <h1>{escape(title)}</h1>
  <div class="meta">
    <div><strong>Filtr:</strong> {escape(filter_label or "bez ograniczenia dat")}</div>
    <div><strong>Liczba wierszy:</strong> {len(rows)}</div>
    <div><strong>Wygenerowano:</strong> {escape(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))}</div>
    {"<ul>" + notes_html + "</ul>" if notes_html else ""}
  </div>
  {chart_html}
  {
    '<table><thead><tr>'
    + ''.join(f'<th>{escape(header)}</th>' for header in headers)
    + '</tr></thead><tbody>'
    + rows_html
    + '</tbody></table>'
    if headers
    else '<div class="empty">Brak danych do eksportu.</div>'
  }
  <div class="footer">Raport wygenerowany z modułu SQL aplikacji Optima SQL GUI.</div>
</body>
</html>
""".strip()


def build_chart_points(headers: list[str], rows: list[dict[str, object]]) -> list[tuple[str, float]]:
    if not headers or not rows:
        return []

    label_header = _find_label_header(headers, rows)
    value_header = _find_numeric_header(headers, rows, exclude=label_header)
    if not label_header or not value_header:
        return []

    aggregated: dict[str, float] = {}
    for row in rows:
        label = str(row.get(label_header, "")).strip()
        if not label:
            continue
        value = _coerce_number(row.get(value_header))
        if value is None:
            continue
        aggregated[label] = aggregated.get(label, 0.0) + value

    ranked = sorted(aggregated.items(), key=lambda item: abs(item[1]), reverse=True)
    return ranked[:8]


def _style_summary_sheet(sheet) -> None:
    fill = PatternFill("solid", fgColor="E7F8F1")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 64


def _style_data_sheet(sheet, headers: list[str], rows: list[dict[str, object]]) -> None:
    fill = PatternFill("solid", fgColor="E7F8F1")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    last_col = get_column_letter(max(len(headers), 1))
    last_row = len(rows) + 1
    if len(headers) >= 1 and len(rows) >= 1:
        table = Table(displayName="RaportSQL", ref=f"A1:{last_col}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    for index, header in enumerate(headers, start=1):
        width = max(len(str(header)), 12)
        for row in rows[:50]:
            width = min(max(width, len(str(row.get(header, ""))) + 2), 40)
        sheet.column_dimensions[get_column_letter(index)].width = width


def _append_chart_sheet(workbook: Workbook, chart_points: list[tuple[str, float]], title: str) -> None:
    chart_sheet = workbook.create_sheet("Wykres")
    chart_sheet.append(["Kategoria", "Wartość"])
    for label, value in chart_points:
        chart_sheet.append([label, value])

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = f"{title} - wykres"
    chart.y_axis.title = "Kategoria"
    chart.x_axis.title = "Wartość"
    data = Reference(chart_sheet, min_col=2, min_row=1, max_row=len(chart_points) + 1)
    categories = Reference(chart_sheet, min_col=1, min_row=2, max_row=len(chart_points) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 16
    chart_sheet.add_chart(chart, "D2")
    chart_sheet.column_dimensions["A"].width = 36
    chart_sheet.column_dimensions["B"].width = 14


def _find_label_header(headers: list[str], rows: list[dict[str, object]]) -> str | None:
    for header in headers:
        values = [str(row.get(header, "")).strip() for row in rows[:50] if str(row.get(header, "")).strip()]
        if not values:
            continue
        if any(_coerce_number(value) is None for value in values):
            return header
    return headers[0] if headers else None


def _find_numeric_header(headers: list[str], rows: list[dict[str, object]], *, exclude: str | None = None) -> str | None:
    best_header = None
    best_score = -1
    for header in headers:
        if header == exclude:
            continue
        score = 0
        for row in rows[:100]:
            if _coerce_number(row.get(header)) is not None:
                score += 1
        if score > best_score:
            best_score = score
            best_header = header
    return best_header if best_score > 0 else None


def _coerce_number(value: object) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isfinite(float(value)):
            return float(value)
        return None

    text = str(value).strip()
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    text = text.replace("\xa0", "").replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    elif re.fullmatch(r"\d{1,3}(\.\d{3})+", text):
        text = text.replace(".", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return -number if negative else number


def _build_chart_svg(chart_points: list[tuple[str, float]]) -> str:
    max_value = max(abs(value) for _, value in chart_points) or 1.0
    row_height = 30
    width = 820
    label_width = 220
    bar_width = 520
    svg_height = 36 + len(chart_points) * row_height
    bars = []
    for index, (label, value) in enumerate(chart_points):
        y = 24 + index * row_height
        normalized = abs(value) / max_value
        current_width = max(6, int(bar_width * normalized))
        bars.append(
            f'<text x="0" y="{y + 14}" font-size="12" fill="#17201c">{escape(label)}</text>'
            f'<rect x="{label_width}" y="{y}" width="{current_width}" height="18" rx="6" fill="#176b5b"></rect>'
            f'<text x="{label_width + current_width + 8}" y="{y + 14}" font-size="12" fill="#17201c">{escape(_format_number(value))}</text>'
        )
    return (
        '<div class="chart"><h2>Wizualizacja</h2>'
        f'<svg width="{width}" height="{svg_height}" viewBox="0 0 {width} {svg_height}" role="img" aria-label="Wykres raportu">'
        + "".join(bars)
        + "</svg></div>"
    )


def _format_number(value: float) -> str:
    return f"{value:,.2f}".replace(",", " ").replace(".", ",")


def _render_pdf_via_browser(file_path: Path, html: str, *, browser_path: str | None = None) -> None:
    browser = browser_path or _detect_browser_path()
    if not browser:
        raise RuntimeError("Brak lokalnej przeglądarki do generowania PDF (Edge/Chrome).")

    file_path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile("w", delete=False, suffix=".html", encoding="utf-8") as handle:
        handle.write(html)
        html_path = Path(handle.name)

    try:
        subprocess.run(
            [
                browser,
                "--headless",
                "--disable-gpu",
                "--allow-file-access-from-files",
                "--print-to-pdf-no-header",
                f"--print-to-pdf={file_path}",
                html_path.resolve().as_uri(),
            ],
            check=True,
            timeout=120,
            capture_output=True,
            text=True,
        )
    except (OSError, subprocess.SubprocessError) as exc:
        raise RuntimeError(f"Nie udało się wygenerować PDF: {exc}") from exc
    finally:
        html_path.unlink(missing_ok=True)

    if not file_path.exists():
        raise RuntimeError("Przeglądarka nie wygenerowała pliku PDF.")


def _detect_browser_path() -> str | None:
    candidates = (
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    )
    for candidate in candidates:
        if Path(candidate).exists():
            return candidate
    return None
