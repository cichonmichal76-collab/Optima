from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

import serve
from src.export.report_sql import build_chart_points, build_sql_report_html, export_sql_report_xlsx


def test_build_chart_points_aggregates_numeric_values():
    headers = ["Kontrahent", "Brutto"]
    rows = [
        {"Kontrahent": "ABC", "Brutto": "100,00"},
        {"Kontrahent": "ABC", "Brutto": "50,00"},
        {"Kontrahent": "XYZ", "Brutto": "25,00"},
    ]

    points = build_chart_points(headers, rows)

    assert points[0] == ("ABC", 150.0)
    assert ("XYZ", 25.0) in points


def test_export_sql_report_xlsx_creates_chart_sheet(tmp_path):
    file_path = tmp_path / "raport.xlsx"
    export_sql_report_xlsx(
        file_path,
        title="Raport SQL",
        headers=["Kontrahent", "Brutto"],
        rows=[
            {"Kontrahent": "ABC", "Brutto": 100},
            {"Kontrahent": "XYZ", "Brutto": 50},
        ],
        notes=["Test eksportu"],
        filter_label="marzec 2026",
        include_chart=True,
    )

    workbook = load_workbook(file_path)
    assert workbook.sheetnames == ["Raport", "Dane", "Wykres"]
    assert workbook["Dane"]["A1"].value == "Kontrahent"
    assert workbook["Raport"]["B2"].value == "Raport SQL"


def test_build_sql_report_html_includes_chart_section():
    html = build_sql_report_html(
        title="Raport SQL",
        headers=["Kontrahent", "Brutto"],
        rows=[{"Kontrahent": "ABC", "Brutto": "120,00"}],
        notes=["Nota 1"],
        filter_label="marzec 2026",
        include_chart=True,
    )

    assert "Wizualizacja" in html
    assert "Raport SQL" in html
    assert "Nota 1" in html


def test_export_report_returns_download_url_for_pdf(monkeypatch, tmp_path):
    monkeypatch.setattr(serve, "ROOT", tmp_path)

    def fake_export_pdf(file_path, **kwargs):
        Path(file_path).write_bytes(b"%PDF-1.4")
        return file_path

    monkeypatch.setattr(serve, "export_sql_report_pdf", fake_export_pdf)

    payload = serve.export_report(
        {
            "format": "pdf",
            "title": "Zestawienie bankowe",
            "headers": ["Opis", "Kwota"],
            "rows": [{"Opis": "Przelew", "Kwota": "10,00"}],
            "include_chart": True,
        }
    )

    assert payload["format"] == "pdf"
    assert payload["include_chart"] is True
    assert payload["download_url"].startswith("/exports/zestawienie-bankowe_wykres_")
